# Copyright (C) 2019 Shymaxtic
# 
# This file is part of ExcelAPI.
# 
# ExcelAPI is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
# 
# ExcelAPI is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
# 
# You should have received a copy of the GNU General Public License
# along with ExcelAPI.  If not, see <http://www.gnu.org/licenses/>.

import openpyxl
from HeaderInfo import HeaderInfo 
from CellInfo import CellInfo
from HeaderInfo import D_SEPERATOR
import Utils

class ExcelFile:

    def __init__(self, path: str):
        self.mPath                  = path    
        self.mHeaderRange           = ""           # Ex: "A1:N1"
        self.mWorkBook              = None
        self.mSheet                 = None
        self.mHeaderList            = {}                # {fullheadername:header info}   
        self.mDataColumnSize        = 0                # number of data column
        self.mDataRowSize           = 0                 # number of data row
        self.mHeaderCellInfo        = {}           # list of info of main cell of header {coordinate:cell info}
        self.mMergedDataCellInfo    = {}        # list of cell in merged cell area {coordinate:cell info}
        self.mPivotRow              = 0                   # row at lowest header.
        self.mPivotColum            = 0                 # colum at first header
        self.mHeaderInfoColumCache  = {}      # {column:fullheadername}
        self.mDictData              = {}   

    def __PostProcessMergedCell(self):
        self.mMergedDataCellInfo, self.mHeaderCellInfo = {}, {}

        for cellRange in self.mSheet.merged_cells.ranges:
            for rowOfCell in self.mSheet[cellRange.__str__()]:
                for cell in rowOfCell:
                    # if this cell is in header range and is top-left cell
                    if Utils.IsCellInCellRange(cell.coordinate, self.mHeaderRange) and \
                        cell == self.mSheet[cellRange.__str__()][0][0]:
                        colNum = len(self.mSheet[cellRange.__str__()][0]) # get column size
                        rowNum = len(self.mSheet[cellRange.__str__()]) # get row size
                        topLeftCell = self.mSheet[cellRange.__str__()][0][0] # get top-left cell
                        self.mHeaderCellInfo[cell.coordinate] = CellInfo(cell, topLeftCell, rowNum, colNum)
                    else:
                        colNum = len(self.mSheet[cellRange.__str__()][0]) # get column size
                        rowNum = len(self.mSheet[cellRange.__str__()]) # get row size
                        topLeftCell = self.mSheet[cellRange.__str__()][0][0] # get top-left cell
                        self.mMergedDataCellInfo[cell.coordinate] = CellInfo(cell, topLeftCell, rowNum, colNum)

        # print(self.mHeaderCellInfo)                

    def __GetHeaderCellInfo(self):
        """
        Get info of cells containing header name, including merged cells or singel cell.
        """
        self.__PostProcessMergedCell()
        # For each row in header sheet    
        for rowOfCell in self.mSheet[self.mHeaderRange]:
            for cell in rowOfCell:
                if cell.value != None and cell.coordinate not in self.mHeaderCellInfo:  # this single cell has value                    
                    self.mHeaderCellInfo[cell.coordinate] = CellInfo(cell, cell, 1, 1)

    def __GetHeaderInfo(self):
        """
        Update relationship of header cells. Create Header Info    
        """
        self.__GetHeaderCellInfo()
        sortedKeys = sorted(self.mHeaderCellInfo.keys())
        self.mHeaderList, self.mHeaderInfoColumCache = {}, {}
        for key in sortedKeys:
            # prepare parent header
            icell = self.mHeaderCellInfo[key].mCell
            if (icell.row > 1):
                # if upper cell (merged cells/single cell) has value. It is parent of this cell
                upperCell = self.mSheet.cell(row=icell.row - 1, column=icell.column)
                # Check if upper cell is in any created header info
                newHeaderInfo = HeaderInfo(self.mHeaderCellInfo[key], None)
                for headerInfo in self.mHeaderList.values():
                    if headerInfo.mCellInfo.Has(upperCell):
                        newHeaderInfo = HeaderInfo(self.mHeaderCellInfo[key], headerInfo)
                        break
                self.mHeaderList[newHeaderInfo.mFullName] = newHeaderInfo
                self.mHeaderInfoColumCache[self.mHeaderCellInfo[key].mCell.column] = newHeaderInfo.mFullName
            else:
                newHeaderInfo = HeaderInfo(self.mHeaderCellInfo[key], None) 
                self.mHeaderList[newHeaderInfo.mFullName] = newHeaderInfo
                self.mHeaderInfoColumCache[self.mHeaderCellInfo[key].mCell.column] = newHeaderInfo.mFullName
        # for key in self.mHeaderList:
        #     print(key, ":::", self.mHeaderList[key])
        # for key in self.mHeaderInfoColumCache:
        #     print(key, ":::", self.mHeaderInfoColumCache[key])                            

    def __CheckMatchHeader(self, checkingName: str, headerFullName: str):
        """
        Check if a name is a header full name. Ex: "Function" is name of header full name "Function:FileName:PIC"
        Note: "Function:" is not name of header full name "Function:FileName:PIC"
        """
        if (checkingName == headerFullName):
            return True
        inputStruct = checkingName.split(D_SEPERATOR, 1)
        checkStruct = headerFullName.split(D_SEPERATOR, 1)
        if (inputStruct[0] == checkStruct[0]):
            if len(inputStruct) > 1 and len(checkStruct) > 1:
                if inputStruct[1] and inputStruct[1] in checkStruct[1]:
                    return True
                else:
                    return False                    
            else:
                return True
        return False  
        
    
    # Get list of matched header with a input name
    # headerStruct: "End Date:Test term"
    # return list of CellInfo (s)
    def __GetMatchHeader(self, headerStruct: str):
        matchHeader = [self.mHeaderList[key].mCellInfo for key in self.mHeaderList if self.__CheckMatchHeader(headerStruct, key)]
        return matchHeader     

    # Get a cell at header name and row offset. Row offset start at 1
    # Ex: GetCell("Function:PIC:Company", 1):
    # return Cell                 
    def GetCell(self, headerStruct: str, rowOffset: int):
        matchHeader = self.__GetMatchHeader(headerStruct)
        if (len(matchHeader) == 0):
                raise Exception("Cannot find header: " +  headerStruct)
        if (len(matchHeader) > 1):
            raise Exception("More than one header: " + headerStruct)
        # print(matchHeader)
        col = matchHeader[0].mCell.column
        row = (matchHeader[0].mCell.row + matchHeader[0].mRowSize - 1) + rowOffset
        # print(row, col)
        # return self.mSheet.cell(row=row, column=col)
        cell = self.mSheet.cell(row=row, column=col)
        if cell.coordinate in self.mMergedDataCellInfo:
            return self.mMergedDataCellInfo[cell.coordinate].mTopLeftCell
        return cell      

   # Load all data and store to mDictData                                
    def __LoadData(self):
        self.mDataRowSize = 0
        rowValue = {}
        rowOffset = 1
        rowValue = self.Read(rowOffset)
        stop = all(value==None for value in rowValue.values())
        # init key for data
        self.mDictData = {key: [] for key in rowValue}
        while (not stop):
            # load each field of row value to mDictData
            # print(rowValue)
            for key in rowValue:
                self.mDictData[key].append(rowValue[key])
            self.mDataRowSize += 1
            rowOffset += 1
            rowValue = self.Read(rowOffset)
            stop = all(value==None for value in rowValue.values())
        # print("self.mDataRowSize=", self.mDataRowSize)                 
        for key in self.mDictData:
            print(key, self.mDictData[key])

    # Write to cell at header name and row offset.
    def Write(self, headerStruct: str, rowOffset: int, value: str):
        cell = self.GetCell(headerStruct, rowOffset)
        cell.value = value          

    def GetValue(self, cell):
        """Get value at a cell with merged lookup
        
        Arguments:
            cell {openpyxl.cell.Cell} -- excel cell
        
        Returns:
            str -- value at cell with merged lookup
        """
        if cell.coordinate in self.mMergedDataCellInfo:
            return self.mMergedDataCellInfo[cell.coordinate].mTopLeftCell.value
        return cell.value            

    def Open(self, sheet=None, headerRange=None, readOnly=False):
        self.mWorkBook = openpyxl.load_workbook(self.mPath, readOnly)
        if (sheet and headerRange):
            self.LoadSheet(sheet, headerRange)

    def LoadSheet(self, sheet:str, headerRange: str):
        self.mSheet = self.mWorkBook[sheet]
        self.mHeaderRange = headerRange
        self.mDataColumnSize = Utils.GetDimension(headerRange)[1]
        self.__GetHeaderInfo()
        self.mPivotRow = openpyxl.utils.cell.coordinate_from_string(self.mHeaderRange.split(":")[1])[1]
        columLetter = openpyxl.utils.cell.coordinate_from_string(self.mHeaderRange.split(":")[0])[0]
        self.mPivotColum = openpyxl.utils.cell.column_index_from_string(columLetter)
        self.__LoadData()

    def Save(self):
        self.mWorkBook.save(self.mPath)
    
    def SaveAs(self, path:str):
        self.mWorkBook.save(path)

    def Read(self, rowOffset: int):
        returnValue = {}
        for i in range(self.mDataColumnSize):
            headerName = self.mHeaderInfoColumCache[self.mPivotColum + i]
            # print(headerName)
            cell = self.mSheet.cell(row=self.mPivotRow+rowOffset,column=self.mPivotColum+i)
            returnValue[headerName] = self.GetValue(cell)
        return returnValue

    def ReadByField(self, headerName: str):
        returnValue = []
        matchHeader = self.__GetMatchHeader(headerName)
        if (len(matchHeader) == 0): raise Exception("Cannot find header: " + headerName)
        if (len(matchHeader) > 1): raise Exception("More than one header: " + headerName)
        # print(matchHeader)
        col = matchHeader[0].mCell.column
        for i in range(1, self.mDataRowSize + 1):
            row = (matchHeader[0].mCell.row + matchHeader[0].mRowSize - 1) + i
            cell = self.mSheet.cell(row=row, column=col)
            value =  self.GetValue(cell)
            returnValue.append(value)
        return returnValue

    # Read data at conditions
    # outputFields: ["C0:Coverage", "C1:Coverage"]
    # conditionFields: {"File Name": "Test.cpp1", "Function Name": "TestFunction"}
    # return {"C0:Coverage" : ["100%"], "C1:Coverage" : ["100%"]} 
    def ReadByCondition(self, outputFields: list, conditionFields):
        # check if output fields is unique
        matchOutputKeys = {}
        outputPairKey = []
        for outputField in outputFields:   # outputField is header name
            tmpKeys = [key for key in self.mDictData if self.__CheckMatchHeader(outputField, key)]
            if (len(tmpKeys) > 1): raise Exception("More than one output field: " + outputField)
            if (len(tmpKeys) == 0): raise Exception("Cannot find output field: " + outputField)                
            matchOutputKeys[tmpKeys[0]] = [] 
            outputPairKey.append((tmpKeys[0], outputField))
        # check if condition field is unique
        matchConditionKeys = {}
        for conditionField in conditionFields: # conditionField is header name
            tmpKeys = [key for key in self.mDictData if self.__CheckMatchHeader(conditionField, key)]
            if (len(tmpKeys) > 1): raise Exception("More than one condition field: " + conditionField)
            if (len(tmpKeys) == 0): raise Exception("Cannot find condition field: " + conditionField)                
            # conditionPairKey.append((tmpKeys[0], conditionField))   
            matchConditionKeys[tmpKeys[0]] = conditionFields[conditionField]        
        # print("matchConditionKeys=", matchConditionKeys)                                                                                            
        # get index of condition field if equal value
        indexs = [i for key in matchConditionKeys for i, val in enumerate(self.mDictData[key]) if val == matchConditionKeys[key]]
        # print(indexs)                
        # get index has more one time
        numOfcond = len(matchConditionKeys)
        matchIdex = set([x for x in indexs if indexs.count(x) == numOfcond])                            
        # print(matchIdex)
        matchOutputKeys = {key: [self.mDictData[key][i] for i in matchIdex] for key in matchOutputKeys}
        # print(matchOutputKeys)  
        returnVal = {pair[1]: matchOutputKeys[pair[0]] for pair in outputPairKey}
        # print(returnVal)                 
        return returnVal

    def ReadRowByCondition(self, conditionFields):
        # check if output fields is unique
        # check if condition field is unique
        matchConditionKeys = {}
        for conditionField in conditionFields: # conditionField is header name
            tmpKeys = [key for key in self.mDictData if self.__CheckMatchHeader(conditionField, key)] # key is header full name
            if (len(tmpKeys) > 1): raise Exception("More than one condition field: " + conditionField)
            if (len(tmpKeys) == 0): raise Exception("Cannot find condition field: " + conditionField)                
            matchConditionKeys[tmpKeys[0]] = conditionFields[conditionField] 
        indexs = [i for key in matchConditionKeys for i, val in enumerate(self.mDictData[key]) if val == matchConditionKeys[key]]
        # print(indexs)                
        # get index has more one time
        numOfcond = len(matchConditionKeys)
        matchIdex = set([x for x in indexs if indexs.count(x) == numOfcond])                            
        # print(matchIdex)

        returnVal = {key:[self.mDictData[key][i] for i in matchIdex] for key in self.mDictData.keys()}
        # print(returnVal)                 
        return returnVal